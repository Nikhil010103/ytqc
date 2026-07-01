"""Phase-2 tests: sinks (ytqc/sinks/) + browser extraction parsing
(ytqc/browser/video_page.py + channel_page.py) driven hermetically by the
FakeKimiClient.

Hermeticity: no network, no real browser, no real clock. The browser tests
patch ytqc.browser.frames.fetch_thumbnail_b64 (the only httpx call reachable
from extract_video) and time.sleep in the frames/transcript modules, and feed
a FakeKimiClient whose screenshot is empty so frame capture resolves to
method=="none" — the parsing assertions are the point, not frame capture.

These tests PIN the QA-audit fixes:
  * F1  — ExcelSink rebuilds from authoritative results.csv (not just this
           process's rows).
  * F2  — extract_video early-returns on an unavailable video (no transcript /
           frame work).
  * F5  — CsvSink.close() dedupes duplicate ids keeping the last.
  * F-injection — to_flat_dict escapes leading formula chars; CsvSink writes the
           escaped cell.
"""
from __future__ import annotations

import csv

import openpyxl
import pytest

from ytqc.models import QCRecord
from ytqc.sinks.base import build_sinks
from ytqc.sinks.csv_sink import CsvSink, COLUMNS as CSV_COLUMNS
from ytqc.sinks.excel_sink import ExcelSink
from ytqc.config import SamplingConfig
from ytqc.browser.video_page import extract_video
from ytqc.browser.channel_page import extract_channel

from tests.fakes import FakeKimiClient
from tests.fixtures import yt_payloads as P


RUN_ID = "run-2026"


def _read_csv_rows(path):
    with open(path, newline="", encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


# ════════════════════════════════════════════════════════════════════════════
# SINKS
# ════════════════════════════════════════════════════════════════════════════

# ── CsvSink: header written once across multiple writes ─────────────────────
def test_csv_header_written_once_across_appends(tmp_path):
    sink = CsvSink()
    sink.open(RUN_ID, str(tmp_path))
    sink.write(QCRecord(id="a", type="video", tier_1="Automobiles"))
    sink.write(QCRecord(id="b", type="video", tier_1="Music"))
    # do NOT close — closing rewrites/dedupes; we want to see the raw append shape
    sink._fh.flush()

    path = tmp_path / RUN_ID / "results.csv"
    text = path.read_text(encoding="utf-8")
    # the header line (starts with the first column name "id,") appears exactly once
    header_line = ",".join(CSV_COLUMNS)
    assert text.count(header_line) == 1
    # two data rows present
    rows = _read_csv_rows(path)
    assert [r["id"] for r in rows] == ["a", "b"]


# ── CsvSink: reopening an existing file does NOT duplicate the header ───────
def test_csv_reopen_does_not_duplicate_header(tmp_path):
    s1 = CsvSink()
    s1.open(RUN_ID, str(tmp_path))
    s1.write(QCRecord(id="a", type="video", tier_1="Automobiles"))
    # close s1 — this dedupes/rewrites and keeps a single header
    s1.close()

    # resume: a fresh sink appends to the same file
    s2 = CsvSink()
    s2.open(RUN_ID, str(tmp_path))
    s2.write(QCRecord(id="b", type="video", tier_1="Music"))
    s2._fh.flush()

    path = tmp_path / RUN_ID / "results.csv"
    text = path.read_text(encoding="utf-8")
    header_line = ",".join(CSV_COLUMNS)
    assert text.count(header_line) == 1


# ── F5: CsvSink.close() dedupes duplicate ids keeping the LAST row ──────────
def test_csv_close_dedupes_duplicate_ids_keep_last(tmp_path):
    sink = CsvSink()
    sink.open(RUN_ID, str(tmp_path))
    # same id "a" written twice with a different tier_1; last should win
    sink.write(QCRecord(id="a", type="video", tier_1="Automobiles"))
    sink.write(QCRecord(id="a", type="video", tier_1="Music"))
    sink.close()

    rows = _read_csv_rows(tmp_path / RUN_ID / "results.csv")
    a_rows = [r for r in rows if r["id"] == "a"]
    assert len(a_rows) == 1, "duplicate id 'a' should be deduped to a single row"
    assert a_rows[0]["tier_1"] == "Music", "dedupe must keep the LAST written value"


# ── F-injection: formula chars in a cell are escaped end-to-end ─────────────
def test_csv_formula_injection_escaped(tmp_path):
    sink = CsvSink()
    sink.open(RUN_ID, str(tmp_path))
    sink.write(QCRecord(id="evil", type="video", name="=HYPERLINK(1)"))
    sink.close()

    rows = _read_csv_rows(tmp_path / RUN_ID / "results.csv")
    cell = rows[0]["name"]
    assert cell.startswith("'"), f"formula cell must be escaped with a leading quote, got {cell!r}"
    assert cell == "'=HYPERLINK(1)"


# ── to_flat_dict directly escapes the other risky leading chars too ─────────
@pytest.mark.parametrize("dangerous", ["=cmd", "+1", "-1+1", "@SUM(1)"])
def test_to_flat_dict_escapes_leading_formula_chars(dangerous):
    rec = QCRecord(id="x", type="video", name=dangerous)
    flat = rec.to_flat_dict()
    assert flat["name"] == "'" + dangerous


# ── build_sinks: name → class mapping, unknown raises KeyError ──────────────
def test_build_sinks_maps_csv_and_xlsx():
    sinks = build_sinks(["csv", "xlsx"])
    assert isinstance(sinks[0], CsvSink)
    assert isinstance(sinks[1], ExcelSink)


def test_build_sinks_unknown_name_raises_keyerror():
    with pytest.raises(KeyError):
        build_sinks(["mongodb"])


# ── F1: ExcelSink rebuilds from authoritative results.csv ───────────────────
def test_excel_rebuilds_all_rows_from_results_csv(tmp_path):
    """Simulate a prior run that wrote 2 rows to results.csv, then a NEW process
    whose ExcelSink only .write()s 1 of them. The produced results.xlsx must
    contain BOTH rows (rebuilt from the authoritative csv), not just the 1."""
    # prior run: write results.csv with 2 distinct ids via the real CsvSink
    csv_sink = CsvSink()
    csv_sink.open(RUN_ID, str(tmp_path))
    csv_sink.write(QCRecord(id="r1", type="video", tier_1="Automobiles",
                            status="OK", brand_safety_is_safe=True,
                            needs_review=False, confidence=0.9))
    csv_sink.write(QCRecord(id="r2", type="video", tier_1="Music",
                            status="OK", brand_safety_is_safe=True,
                            needs_review=False, confidence=0.9))
    csv_sink.close()

    # new process: ExcelSink writes only ONE row in-memory, then closes
    xl = ExcelSink()
    xl.open(RUN_ID, str(tmp_path))
    xl.write(QCRecord(id="r1", type="video", tier_1="Automobiles",
                      status="OK", brand_safety_is_safe=True,
                      needs_review=False, confidence=0.9))
    xl.close()

    xlsx_path = tmp_path / RUN_ID / "results.xlsx"
    assert xlsx_path.exists()
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    # header row + 2 data rows
    assert ws.max_row == 3, "xlsx must rebuild BOTH csv rows, not just the one written this process"
    id_col = CSV_COLUMNS.index("id") + 1
    ids = {ws.cell(row=r, column=id_col).value for r in range(2, ws.max_row + 1)}
    assert ids == {"r1", "r2"}


# ── ExcelSink conditional fills: green safe, red unsafe, grey error ─────────
def test_excel_conditional_fill_colors(tmp_path):
    # CsvSink is authoritative; write 3 representative rows through it.
    csv_sink = CsvSink()
    csv_sink.open(RUN_ID, str(tmp_path))
    # green: safe + confident (status OK, is_safe True, needs_review False)
    csv_sink.write(QCRecord(id="green", type="video", status="OK",
                            brand_safety_is_safe=True, needs_review=False))
    # red: brand_safety_is_safe False
    csv_sink.write(QCRecord(id="red", type="video", status="OK",
                            brand_safety_is_safe=False, needs_review=False))
    # grey: status ERROR
    csv_sink.write(QCRecord(id="err", type="video", status="ERROR",
                            error="boom", brand_safety_is_safe=None))
    csv_sink.close()

    xl = ExcelSink()
    xl.open(RUN_ID, str(tmp_path))
    xl.close()

    wb = openpyxl.load_workbook(tmp_path / RUN_ID / "results.xlsx")
    ws = wb.active
    id_col = CSV_COLUMNS.index("id") + 1
    # map id -> the row's fill color (read off the id cell; whole row shares fill)
    color_by_id = {}
    for r in range(2, ws.max_row + 1):
        rid = ws.cell(row=r, column=id_col).value
        color_by_id[rid] = ws.cell(row=r, column=id_col).fill.fgColor.rgb

    def rgb(hex6):
        # openpyxl stores solid fills as 8-hex (AARRGGBB)
        return "00" + hex6

    assert color_by_id["green"] == rgb("C6EFCE")
    assert color_by_id["red"] == rgb("FFC7CE")
    assert color_by_id["err"] == rgb("D9D9D9")


# ════════════════════════════════════════════════════════════════════════════
# BROWSER — extract_video / extract_channel via FakeKimiClient
# ════════════════════════════════════════════════════════════════════════════

@pytest.fixture
def no_network_frames(monkeypatch):
    """Neutralise the only network call reachable from extract_video
    (thumbnail fetch) and the real clock used by frame/transcript waits."""
    import ytqc.browser.frames as frames_mod
    import ytqc.browser.transcript as transcript_mod
    monkeypatch.setattr(frames_mod, "fetch_thumbnail_b64", lambda video_id: None)
    monkeypatch.setattr(frames_mod.time, "sleep", lambda *a, **k: None)
    monkeypatch.setattr(transcript_mod.time, "sleep", lambda *a, **k: None)
    return monkeypatch


def _video_kimi(player_response, **extra_routes):
    """A FakeKimiClient that returns empty transcript/frame data so frame
    capture resolves to method=='none' and parsing is the only thing exercised.
    screenshot='' keeps the screenshot-fallback frame out of the FrameSet."""
    routes = {
        "player_response": player_response,
        "likes": P.LIKES_OK,
        "transcript_open": P.TRANSCRIPT_OPEN_NO_BUTTON,
        "transcript_scrape": P.TRANSCRIPT_SCRAPE_EMPTY,
        "comments": P.COMMENTS_OK,
        # Report each seek instantly "ready" so the frame seek-poll breaks on its
        # first iteration. Without this, the bounded `while time.time() < deadline`
        # poll (6s/frame) becomes a real-wall-clock busy-loop once time.sleep is
        # no-op'd — 5 frames -> ~30s. The canvas grab still fails (default ok:False)
        # and screenshot is "" so frames.method stays "none" (the test intent).
        "frame_ready": {"seeking": False, "ready": True, "ad": False},
        "ad_skip": P.AD_SKIP_NONE,
    }
    routes.update(extra_routes)
    # default {"ok": False} for unrouted JS (frame grab etc.) → canvas dead;
    # screenshot "" → no screenshot frame → frames.method == "none".
    return FakeKimiClient(routes, default={"ok": False}, screenshot="")


def test_extract_video_ok_maps_player_response_fields(no_network_frames):
    kimi = _video_kimi(P.PLAYER_RESPONSE_OK)
    ex = extract_video(kimi, "vid123", SamplingConfig(), with_comments=False)

    assert ex.ok is True
    assert ex.title == "Kawasaki Ninja ZX-4R Track Day Review"
    assert ex.author == "MotoGarage"
    assert ex.channel_id == "UC1234567890abcdefABCD12"
    assert ex.duration_s == 814.0
    assert ex.view_count == 100711
    assert ex.youtube_category == "Autos & Vehicles"
    assert ex.is_family_safe is True
    assert ex.is_live is False
    assert ex.keywords == ["kawasaki", "ninja zx-4r", "track day", "motorcycle review"]
    assert ex.provenance["metadata"] == "player_response"
    # navigated to the watch page
    assert any("watch?v=vid123" in u for u in kimi.navigated)


def test_extract_video_ok_views_per_day_computed(no_network_frames):
    kimi = _video_kimi(P.PLAYER_RESPONSE_OK)
    ex = extract_video(kimi, "vid123", SamplingConfig(), with_comments=False)
    # publishDate is recent (2026-05-12) vs today; views_per_day is finite + > 0
    assert ex.days_since_publish > 0
    assert ex.views_per_day > 0
    # views_per_day == round(view_count / max(days, 1.0), 1)
    expected = round(ex.view_count / max(ex.days_since_publish, 1.0), 1)
    assert ex.views_per_day == expected


def test_extract_video_unavailable_returns_early(no_network_frames):
    """F2: an unavailable/deleted video must bail before any transcript/frame
    work — no hallucinated verdict on empty details."""
    kimi = _video_kimi(P.PLAYER_RESPONSE_UNAVAILABLE)
    ex = extract_video(kimi, "deadvid", SamplingConfig(), with_comments=True)

    assert ex.ok is False
    assert ex.error  # reason / status set
    assert ex.provenance["metadata"] == "unavailable"
    # early return: provenance never advanced to transcript/frames/likes
    assert "transcript" not in ex.provenance
    assert "frames" not in ex.provenance
    assert "likes" not in ex.provenance
    # and crucially the fake was never asked for transcript / frame / likes JS
    from ytqc.browser import youtube_js as J
    joined = "\n".join(kimi.js_calls)
    assert J.TRANSCRIPT_OPEN not in kimi.js_calls
    assert J.LIKES not in kimi.js_calls
    assert J.FRAME_GRAB not in joined
    # frame capture never ran → FrameSet stayed at its default method
    assert ex.frames.method == "none"


def test_extract_video_no_player_response_fails(no_network_frames):
    kimi = _video_kimi(P.PLAYER_RESPONSE_NO_PR)
    ex = extract_video(kimi, "nopr", SamplingConfig(), with_comments=False)
    assert ex.ok is False
    assert ex.error == "player response unavailable (page failed to load?)"
    assert ex.provenance["metadata"] == "none"


def test_extract_video_bad_date_falls_back_to_365(no_network_frames):
    """Empty/non-ISO publishDate → days_since_publish keeps its 365.0 default
    and views_per_day is computed without crashing."""
    kimi = _video_kimi(P.PLAYER_RESPONSE_BAD_DATE)
    ex = extract_video(kimi, "baddate", SamplingConfig(), with_comments=False)
    assert ex.ok is True
    assert ex.days_since_publish == 365.0
    # viewCount 36500 / 365 == 100.0
    assert ex.views_per_day == round(36500 / 365.0, 1)


def test_extract_video_non_iso_date_falls_back_to_365(no_network_frames):
    kimi = _video_kimi(P.PLAYER_RESPONSE_BAD_DATE_NONISO)
    ex = extract_video(kimi, "baddate2", SamplingConfig(), with_comments=False)
    assert ex.ok is True
    assert ex.days_since_publish == 365.0
    assert ex.views_per_day == round(36500 / 365.0, 1)


def test_extract_video_frames_method_none_when_no_frame_data(no_network_frames):
    """With FRAME_GRAB unrouted (default {ok:False}) and screenshot empty, frame
    capture yields no usable frame → method == 'none' and the parse still
    succeeds."""
    kimi = _video_kimi(P.PLAYER_RESPONSE_OK)
    ex = extract_video(kimi, "vidframes", SamplingConfig(), with_comments=False)
    assert ex.ok is True
    assert ex.frames.method == "none"
    assert ex.frames.frames_b64 == []


# ── extract_channel ─────────────────────────────────────────────────────────
def test_extract_channel_parses_about_and_videos(no_network_frames):
    kimi = FakeKimiClient(
        {
            "channel_about": P.CHANNEL_ABOUT_OK,
            "channel_videos": P.CHANNEL_VIDEOS_OK,
        },
        default={"ok": False},
        screenshot=P.TINY_JPEG_B64,
    )
    ex = extract_channel(kimi, "UC1234567890abcdefABCD12")

    assert ex.ok is True
    assert ex.title == "MotoGarage"
    assert ex.external_id == "UC1234567890abcdefABCD12"
    assert ex.subscribers == 1_200_000          # "1.2M subscribers"
    assert ex.total_views == 458_000_000         # "458,000,000 views"
    assert ex.video_count == 642                 # "642 videos"
    assert ex.country == "United States"
    assert ex.joined_date == "Mar 3, 2015"       # "Joined " stripped
    assert ex.provenance["about"] == "ytInitialData"
    # fake can't run the async continuation API → graceful fallback to lockupViewModel
    assert ex.provenance["videos_grid"] == "lockupViewModel:6"

    # recent_videos populated (6 tiles) with parsed views/days
    assert len(ex.recent_videos) == 6
    first = ex.recent_videos[0]
    assert first.video_id == "vid0001aaaa"
    assert first.views == 100_000                # "100K views"
    assert first.days_ago == 2.0                 # "2 days ago"

    # navigated to both /about and /videos
    assert any(u.endswith("/about") for u in kimi.navigated)
    assert any(u.endswith("/videos") for u in kimi.navigated)


def test_extract_channel_velocity_computed(no_network_frames):
    kimi = FakeKimiClient(
        {"channel_about": P.CHANNEL_ABOUT_OK, "channel_videos": P.CHANNEL_VIDEOS_OK},
        default={"ok": False},
        screenshot=P.TINY_JPEG_B64,
    )
    ex = extract_channel(kimi, "UC1234567890abcdefABCD12")

    last5 = [v.views for v in ex.recent_videos[:5]]
    prev5 = [v.views for v in ex.recent_videos[5:10]]
    avg_r = sum(last5) / len(last5)
    avg_p = sum(prev5) / len(prev5) if prev5 else 0
    assert ex.avg_views_last5 == round(avg_r, 0)
    assert ex.avg_views_prev5 == round(avg_p, 0)
    assert ex.velocity_score == round((avg_r - avg_p) / max(avg_p, 1), 4)
