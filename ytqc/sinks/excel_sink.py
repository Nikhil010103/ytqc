"""Styled Excel sink — styling lifted from yt_qc_checker.save_output, with
the recolor rule: green = safe+confident, amber = needs_review, red = unsafe,
grey = error."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ytqc.models import QCRecord
from ytqc.sinks.base import ResultSink

COLUMNS = list(QCRecord.model_fields.keys())

_WIDE = {"name": 44, "summary": 60, "comment": 44, "tier_classification_reasoning": 50,
         "brand_safety_explanation": 50, "keywords": 36, "lookalike_keywords": 36, "topics": 36,
         "content_themes": 30, "audience_interests": 30,
         "brand_safety_triggered_categories": 30, "error": 40,
         "vidiq_insight": 60, "vidiq_signals": 44, "vidiq_similar_channels": 30,
         "vidiq_controversial_keywords": 30}


class ExcelSink(ResultSink):
    def __init__(self):
        self._rows: list[dict] = []
        self.path: Path | None = None
        self._run_dir: Path | None = None

    def open(self, run_id: str, output_dir: str) -> None:
        self._run_dir = Path(output_dir) / run_id
        self.path = self._run_dir / "results.xlsx"
        self.path.parent.mkdir(parents=True, exist_ok=True)

    def write(self, record: QCRecord) -> None:
        self._rows.append(record.to_flat_dict())

    def _collect_rows(self) -> pd.DataFrame | None:
        """Resume-safe row source.

        results.csv (written by CsvSink, flushed per row and closed before this
        sink in cli ordering) is the authoritative record of ALL rows across all
        resume attempts. Prefer it. Its list columns are already '; '-joined
        strings and bool fields are "True"/"False"/"ERROR" strings, so the
        styling logic below still works unchanged.

        If results.csv is absent (xlsx-only run), fall back to merging any prior
        results.xlsx with this process's rows, deduping by id (keep last).
        """
        csv_path = self._run_dir / "results.csv" if self._run_dir else None
        if csv_path is not None and csv_path.exists():
            df = pd.read_csv(csv_path, dtype=str, keep_default_na=False)
            if df.empty:
                return None
            return df.reindex(columns=COLUMNS)

        # xlsx-only fallback: merge prior workbook rows with current process rows.
        prior: list[dict] = []
        if self.path is not None and self.path.exists():
            prev = pd.read_excel(self.path, engine="openpyxl", dtype=str)
            prev = prev.fillna("")
            prior = prev.to_dict("records")
        combined = prior + self._rows
        if not combined:
            return None
        df = pd.DataFrame(combined).reindex(columns=COLUMNS)
        if "id" in df.columns:
            df = df.drop_duplicates(subset="id", keep="last")
        return df

    def close(self) -> None:
        if self.path is None:
            return
        df = self._collect_rows()
        if df is None or df.empty:
            return
        df = df.reindex(columns=COLUMNS)
        df.to_excel(self.path, index=False, engine="openpyxl")

        wb = openpyxl.load_workbook(self.path)
        ws = wb.active

        header_fill = PatternFill("solid", fgColor="1A1A2E")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        safe_fill = PatternFill("solid", fgColor="C6EFCE")     # green
        review_fill = PatternFill("solid", fgColor="FFEB9C")   # amber
        unsafe_fill = PatternFill("solid", fgColor="FFC7CE")   # red
        error_fill = PatternFill("solid", fgColor="D9D9D9")    # grey
        thin = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin
        ws.row_dimensions[1].height = 28

        idx = {c: i for i, c in enumerate(COLUMNS)}
        for row in ws.iter_rows(min_row=2):
            status = row[idx["status"]].value
            is_safe = str(row[idx["brand_safety_is_safe"]].value)
            needs_review = str(row[idx["needs_review"]].value)
            if status == "ERROR":
                base = error_fill
            elif is_safe == "False":
                base = unsafe_fill
            elif needs_review == "True":
                base = review_fill
            else:
                base = safe_fill
            for cell in row:
                cell.border = thin
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.fill = base

        for i, col in enumerate(COLUMNS, 1):
            ws.column_dimensions[get_column_letter(i)].width = _WIDE.get(col, 15)

        ws.freeze_panes = "D2"
        wb.save(self.path)
